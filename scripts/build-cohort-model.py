#!/usr/bin/env python3
"""
RegGuard AI — 36-Month Cohort Financial Model
==============================================

Expands §4.2 unit economics from the Business Assessment PDF into a
36-month cohort model showing:

  - ARR growth (new / expansion / churn / net new / cumulative)
  - Monthly burn (ramps as headcount grows per Use-of-Funds §5.2)
  - Cash balance (starts at $2.5M seed close)
  - Runway (months remaining at current burn)
  - Cohort revenue decay (each cohort churns over 5-year lifetime)

Three acquisition scenarios on a single Assumptions sheet:
  - Conservative:  6 new customers in Y1, ramping to 24/yr by Y3
  - Base:         12 new customers in Y1, ramping to 48/yr by Y3
  - Aggressive:   18 new customers in Y1, ramping to 72/yr by Y3

Output: /home/z/my-project/download/RegGuard_AI_36Month_Cohort_Model.xlsx
"""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList

# ─────────────────────────────────────────────────────────────
# Design tokens (dark editorial palette matching the Business Assessment PDF)
# ─────────────────────────────────────────────────────────────

PRIMARY      = '0F2942'   # deep navy
ACCENT       = 'C2410C'   # burnt orange
ACCENT_GREEN = '166534'   # forest green (positive)
ACCENT_RED   = '991B1B'   # oxblood (negative)
ACCENT_GOLD  = 'B45309'   # gold (caution)
BG_ASSUMPTION = 'FEF3C7'  # soft yellow
BG_HEADER    = '0F2942'   # navy header
BG_SUBHEADER = '1E3A5F'   # mid navy
BG_ROW_ALT   = 'F1F5F9'   # zebra stripe
BG_TOTAL     = 'DBEAFE'   # totals row

FONT_BODY     = 'Calibri'
FONT_HEADING  = 'Calibri'

# ─────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────

def f_body(size=10, bold=False, color='1F2937', italic=False):
    return Font(name=FONT_BODY, size=size, bold=bold, color=color, italic=italic)

def f_heading(size=12, color='FFFFFF'):
    return Font(name=FONT_HEADING, size=size, bold=True, color=color)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin(color='CBD5E1'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_top(color='0F2942', style='thin'):
    return Border(top=Side(style=style, color=color))

# Number formats (per xlsx finance scene rules)
FMT_CCY_K   = '$#,##0,"K";($#,##0,"K");"-"'
FMT_CCY_M   = '$#,##0.0,,"M";($#,##0.0,,"M");"-"'
FMT_CCY_PLAIN = '$#,##0;($#,##0);"-"'
FMT_PCT     = '0.0%;(0.0%);"-"'
FMT_INT     = '#,##0;(#,##0);"-"'
FMT_DECIMAL = '#,##0.00;(#,##0.00);"-"'
FMT_MONTHS  = '0.0" mo";(0.0" mo");"-"'

# ─────────────────────────────────────────────────────────────
# Assumptions (from Business Assessment §4.2 + §5.2)
# ─────────────────────────────────────────────────────────────

ASSUMPTIONS = {
    # Unit economics (§4.2)
    'acv_usd':              70_000,    # Average contract value
    'gross_margin':         0.78,      # LLM API + connector infra = 22% COGS
    'customer_lifetime_yr': 5.0,       # Mid-market compliance tools
    'nrr_annual':           1.15,      # 115% net revenue retention
    'cac_usd':              28_000,    # Blended direct + channel
    'ltv_usd':              385_000,   # ACV × lifetime × NRR
    'ltv_cac_ratio':        13.8,      # 385 / 28
    'cac_payback_months':   7.2,       # CAC / (ACV × GM / 12)

    # Funding (§5.1, §5.2)
    'seed_raise_usd':       2_500_000,
    'seed_close_month':     1,         # Month 1 of model = seed close
    'target_series_a_arr':  1_200_000, # Series A trigger
    'series_a_raise_usd':   12_500_000,# Target $10-15M
    'series_a_month_target': 18,       # Target close month

    # Burn schedule (Use of Funds §5.2 — 55% eng / 25% GTM / 10% security / 10% ops)
    # Total monthly burn ramps as headcount grows:
    #   Months 1-6:   founders + 1 eng + 1 GTM ≈ $140K/mo
    #   Months 7-12:  +2 eng + 1 GTM            ≈ $180K/mo
    #   Months 13-18: +1 ML/infra               ≈ $220K/mo (peak pre-A)
    #   Months 19-24: post-A expansion          ≈ $310K/mo (4 new hires)
    #   Months 25-30: scale GTM                 ≈ $420K/mo
    #   Months 31-36: full team                  ≈ $560K/mo
    'burn_schedule': [
        (1, 6,   140_000),
        (7, 12,  180_000),
        (13, 18, 220_000),
        (19, 24, 310_000),
        (25, 30, 420_000),
        (31, 36, 560_000),
    ],

    # Direct vs channel mix (§4.2: "60% direct, 40% channel")
    'direct_mix':           0.60,
    'channel_mix':          0.40,
    'cac_direct':           35_000,
    'cac_channel':          18_000,

    # Scenarios: new customers per year, by year (Y1 / Y2 / Y3)
    'scenarios': {
        'Conservative': {'y1': 6,  'y2': 14, 'y3': 24, 'color': '991B1B'},
        'Base':         {'y1': 12, 'y2': 28, 'y3': 48, 'color': '0F2942'},
        'Aggressive':   {'y1': 18, 'y2': 42, 'y3': 72, 'color': '166534'},
    },
}

# Derived constants
MONTHLY_CHURN = 1.0 / (ASSUMPTIONS['customer_lifetime_yr'] * 12)  # 1/60 = 1.67%/mo
MONTHLY_EXPANSION = (ASSUMPTIONS['nrr_annual'] ** (1/12)) - 1     # ~1.17%/mo
TOTAL_MONTHS = 36

# ─────────────────────────────────────────────────────────────
# Model computation
# ─────────────────────────────────────────────────────────────

def customers_per_month(scenario_yearly):
    """Spread yearly new-customer count evenly across 12 months of that year."""
    monthly = []
    for y in ['y1', 'y2', 'y3']:
        per_month = scenario_yearly[y] / 12.0
        for _ in range(12):
            monthly.append(per_month)
    return monthly  # length 36

def burn_for_month(month_idx):
    """Return monthly burn for month_idx (1-indexed)."""
    for start, end, burn in ASSUMPTIONS['burn_schedule']:
        if start <= month_idx <= end:
            return burn
    return ASSUMPTIONS['burn_schedule'][-1][2]

def simulate_scenario(scenario_name, scenario_yearly):
    """Run 36-month cohort simulation. Returns list of monthly dicts."""
    new_per_month = customers_per_month(scenario_yearly)
    months = []

    # Cohort tracker: each cohort enters with N customers paying ACV/12 per month
    # Cohort revenue = active_customers_in_cohort × ACV/12
    # Active customers decay by MONTHLY_CHURN each month
    # Expansion: each surviving customer pays (1 + MONTHLY_EXPANSION)^(months_since_signed)
    cohorts = []  # list of {sign_month, initial_customers, current_customers, arr_per_customer_at_sign}

    cash_balance = ASSUMPTIONS['seed_raise_usd']
    series_a_received = False
    cumulative_arr = 0.0
    cumulative_new_customers = 0.0
    cumulative_churned_customers = 0.0

    for m in range(1, TOTAL_MONTHS + 1):
        # 1. New customers this month
        new_cust = new_per_month[m - 1]

        # 2. Add new cohort
        cohorts.append({
            'sign_month': m,
            'initial_customers': new_cust,
            'current_customers': new_cust,
            'arr_per_customer_at_sign': ASSUMPTIONS['acv_usd'],
        })
        cumulative_new_customers += new_cust

        # 3. Decay existing cohorts (apply churn BEFORE counting this month's revenue)
        #    (Cohorts that signed this month are not yet churned)
        churned_this_month = 0.0
        for c in cohorts[:-1]:  # skip the one we just added
            churned = c['current_customers'] * MONTHLY_CHURN
            c['current_customers'] -= churned
            churned_this_month += churned

        cumulative_churned_customers += churned_this_month

        # 4. Compute revenue components
        # New ARR: cohorts signed THIS month × ACV
        new_arr = new_cust * ASSUMPTIONS['acv_usd']

        # Churned ARR: customers that churned this month × their current per-customer ARR
        # (current per-customer ARR includes expansion accumulated)
        churned_arr = 0.0
        for c in cohorts[:-1]:
            months_since_sign = m - c['sign_month']
            current_per_customer_arr = c['arr_per_customer_at_sign'] * (
                (1 + MONTHLY_EXPANSION) ** months_since_sign
            )
            churned_arr += (churned_this_month / max(c['current_customers'] + churned_this_month * 1.0001, 0.0001)) * 0  # avoid div0; we'll compute differently below

        # Simpler churn-ARR calc: total churned customers × average current ACV
        avg_current_acv = ASSUMPTIONS['acv_usd'] * ((1 + MONTHLY_EXPANSION) ** (m / 2))  # approx
        churned_arr = churned_this_month * avg_current_acv

        # Expansion ARR: surviving customers × monthly expansion × their current ACV
        expansion_arr = 0.0
        for c in cohorts:
            months_since_sign = m - c['sign_month']
            if months_since_sign == 0:
                continue  # brand new cohort, no expansion yet
            # Expansion = current_customers × ACV × ((1+e)^m - (1+e)^(m-1))
            prev_multiplier = (1 + MONTHLY_EXPANSION) ** (months_since_sign - 1)
            curr_multiplier = (1 + MONTHLY_EXPANSION) ** months_since_sign
            delta_per_customer = c['arr_per_customer_at_sign'] * (curr_multiplier - prev_multiplier)
            expansion_arr += c['current_customers'] * delta_per_customer

        # Total active customers
        active_customers = sum(c['current_customers'] for c in cohorts)

        # Cumulative ARR = sum across cohorts of (current_customers × current_per_customer_arr)
        total_arr = 0.0
        for c in cohorts:
            months_since_sign = m - c['sign_month']
            current_per_customer_arr = c['arr_per_customer_at_sign'] * (
                (1 + MONTHLY_EXPANSION) ** months_since_sign
            )
            total_arr += c['current_customers'] * current_per_customer_arr
        cumulative_arr = total_arr  # cumulative_arr IS total_arr at month m

        # Net new ARR = new + expansion - churn
        net_new_arr = new_arr + expansion_arr - churned_arr

        # 5. Cash flow
        monthly_revenue = cumulative_arr / 12.0  # ARR / 12 = monthly recognized revenue
        gross_profit = monthly_revenue * ASSUMPTIONS['gross_margin']
        burn = burn_for_month(m)

        # Series A trigger: when cumulative_arr >= $1.2M AND month >= 15
        series_a_cash_inflow = 0
        if (not series_a_received
            and cumulative_arr >= ASSUMPTIONS['target_series_a_arr']
            and m >= 15):
            series_a_received = True
            series_a_cash_inflow = ASSUMPTIONS['series_a_raise_usd']

        net_cash_flow = gross_profit - burn + series_a_cash_inflow
        cash_balance += net_cash_flow

        # Runway: months remaining at current burn (using gross profit net of burn)
        # If gross profit > burn, runway = infinite (mark as "Profitable")
        net_burn = max(0, burn - gross_profit)
        if net_burn == 0:
            runway_months = float('inf')
        else:
            runway_months = cash_balance / net_burn if cash_balance > 0 else 0

        months.append({
            'month': m,
            'new_customers': new_cust,
            'churned_customers': churned_this_month,
            'active_customers': active_customers,
            'new_arr': new_arr,
            'expansion_arr': expansion_arr,
            'churned_arr': churned_arr,
            'net_new_arr': net_new_arr,
            'cumulative_arr': cumulative_arr,
            'gross_profit': gross_profit,
            'burn': burn,
            'series_a_cash': series_a_cash_inflow,
            'net_cash_flow': net_cash_flow,
            'cash_balance': cash_balance,
            'runway_months': runway_months,
            'series_a_received': series_a_received,
        })

    return months


# ─────────────────────────────────────────────────────────────
# Workbook builder
# ─────────────────────────────────────────────────────────────

def style_header_row(ws, row, start_col, end_col, fill_color=BG_HEADER, font_color='FFFFFF', height=22):
    ws.row_dimensions[row].height = height
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(fill_color)
        cell.font = f_heading(size=10, color=font_color)
        cell.alignment = align('center', 'center')
        cell.border = border_thin()

def style_assumption_cell(cell, comment_text=None):
    cell.font = Font(name=FONT_BODY, size=10, color='0000FF', bold=False)  # blue = input
    cell.fill = fill(BG_ASSUMPTION)
    cell.alignment = align('right', 'center')
    cell.border = border_thin()
    if comment_text:
        cell.comment = Comment(comment_text, 'Z.ai')

def style_derived_cell(cell):
    cell.font = Font(name=FONT_BODY, size=10, color='166534', italic=True)  # green = derived
    cell.alignment = align('right', 'center')
    cell.border = border_thin()

def style_total_row(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = f_body(bold=True, color=PRIMARY)
        cell.fill = fill(BG_TOTAL)
        cell.border = border_top(color=PRIMARY, style='medium')

# ─────────────────────────────────────────────────────────────
# Sheet 1: README
# ─────────────────────────────────────────────────────────────

def build_readme(wb):
    ws = wb.create_sheet('README', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 100

    # Title
    ws.merge_cells('B2:B2')
    ws['B2'] = 'RegGuard AI — 36-Month Cohort Financial Model'
    ws['B2'].font = Font(name=FONT_HEADING, size=18, bold=True, color=PRIMARY)
    ws['B2'].alignment = align('left', 'center')
    ws.row_dimensions[2].height = 30

    ws['B3'] = 'Source: §4.2 Unit Economics + §5.2 Use of Funds, Business Assessment (Aug 2026)'
    ws['B3'].font = f_body(size=10, italic=True, color='6B7280')

    # Sections
    rows = [
        ('Sheet',         'Contents'),
        ('README',        'This sheet — model overview, color key, and reading order.'),
        ('Assumptions',   'All inputs (blue font, yellow bg). Edit these to flex the model. Derived values in green italic.'),
        ('Conservative',  '6 → 14 → 24 new customers per year. Slow design-partner ramp.'),
        ('Base',          '12 → 28 → 48 new customers per year. The plan we pitch to investors.'),
        ('Aggressive',    '18 → 42 → 72 new customers per year. Best-case if MiCA + Consumer Duty demand pulls forward.'),
        ('Summary',       'Side-by-side comparison: ARR at M12/M24/M36, runway, peak burn, Series A timing.'),
    ]
    start_row = 5
    for i, (sheet, contents) in enumerate(rows):
        r = start_row + i
        ws.cell(row=r, column=2, value=sheet)
        ws.cell(row=r, column=3, value=contents)
        ws.column_dimensions['C'].width = 90
        if i == 0:
            ws.cell(row=r, column=2).font = f_heading(size=11)
            ws.cell(row=r, column=2).fill = fill(BG_HEADER)
            ws.cell(row=r, column=3).font = f_heading(size=11)
            ws.cell(row=r, column=3).fill = fill(BG_HEADER)
        else:
            ws.cell(row=r, column=2).font = f_body(bold=True, color=PRIMARY)
            ws.cell(row=r, column=3).font = f_body()
            ws.cell(row=r, column=2).alignment = align('left', 'center')
            ws.cell(row=r, column=3).alignment = align('left', 'center', wrap=True)
            if i % 2 == 0:
                ws.cell(row=r, column=2).fill = fill(BG_ROW_ALT)
                ws.cell(row=r, column=3).fill = fill(BG_ROW_ALT)
            ws.cell(row=r, column=2).border = border_thin()
            ws.cell(row=r, column=3).border = border_thin()
        ws.row_dimensions[r].height = 20

    # Color key
    key_start = start_row + len(rows) + 2
    ws.cell(row=key_start, column=2, value='Color Key').font = f_body(bold=True, size=12, color=PRIMARY)
    key_items = [
        ('Blue font on yellow bg', 'Hardcoded input assumption — safe to edit'),
        ('Green italic font',       'Derived value computed from inputs — do not edit'),
        ('Black font',              'Computed cell (formula or model output) — do not edit'),
        ('Red font',                'Negative value (cash outflow, churn, deficit)'),
        ('Bold navy on blue bg',    'Total / subtotal row'),
    ]
    for i, (label, desc) in enumerate(key_items):
        r = key_start + 1 + i
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT_BODY, size=10, color='0000FF')
        ws.cell(row=r, column=2).fill = fill(BG_ASSUMPTION)
        ws.cell(row=r, column=3, value=desc).font = f_body()
        ws.cell(row=r, column=2).border = border_thin()
        ws.cell(row=r, column=3).border = border_thin()

    # Methodology
    meth_start = key_start + len(key_items) + 3
    ws.cell(row=meth_start, column=2, value='Methodology').font = f_body(bold=True, size=12, color=PRIMARY)
    meth_items = [
        'Cohort model: each month\'s new customers form a cohort. Cohorts decay at the monthly churn rate (1 / (lifetime × 12) = 1.67%/mo).',
        'Expansion revenue: each surviving customer\'s ACV grows by (NRR^(1/12) − 1) per month ≈ 1.17%/mo, reflecting plugin upsell + connector expansion.',
        'Burn schedule follows §5.2 Use of Funds: 55% engineering, 25% GTM, 10% security, 10% ops. Burn ramps as headcount grows (4 hires in Y1, 2 more pre-A, post-A expansion).',
        'Series A trigger: when cumulative ARR ≥ $1.2M AND month ≥ 15, the model injects $12.5M (target $10-15M) into cash balance.',
        'Runway = cash balance / max(0, burn − gross profit). When gross profit ≥ burn, runway is infinite (model shows "Profitable").',
        'Churned ARR is approximated as churned_customers × average current ACV (blended across all cohorts). This is conservative; actual churn-ARR will be slightly lower because older cohorts have lower ACV.',
        'CAC is paid upfront in the month of acquisition (deducted from cash via burn — not modeled separately to avoid double-counting with the burn schedule).',
    ]
    for i, item in enumerate(meth_items):
        r = meth_start + 1 + i
        ws.cell(row=r, column=2, value=f'• {item}').font = f_body(size=10)
        ws.cell(row=r, column=2).alignment = align('left', 'top', wrap=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 32

    # Footer
    foot_row = meth_start + len(meth_items) + 3
    ws.cell(row=foot_row, column=2, value='Generated by RegGuard AI — Aug 2026').font = f_body(size=9, italic=True, color='6B7280')


# ─────────────────────────────────────────────────────────────
# Sheet 2: Assumptions
# ─────────────────────────────────────────────────────────────

def build_assumptions(wb):
    ws = wb.create_sheet('Assumptions')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 60

    # Title
    ws['B2'] = 'Model Assumptions'
    ws['B2'].font = Font(name=FONT_HEADING, size=16, bold=True, color=PRIMARY)
    ws.row_dimensions[2].height = 28
    ws['B3'] = 'All blue/yellow cells are editable inputs. Green italic cells are derived.'
    ws['B3'].font = f_body(size=10, italic=True, color='6B7280')

    row = 5
    # Section header
    ws.merge_cells(f'B{row}:D{row}')
    ws.cell(row=row, column=2, value='Unit Economics (§4.2)')
    style_header_row(ws, row, 2, 4, fill_color=BG_SUBHEADER)
    row += 1
    # Headers
    ws.cell(row=row, column=2, value='Input')
    ws.cell(row=row, column=3, value='Value')
    ws.cell(row=row, column=4, value='Source / Notes')
    style_header_row(ws, row, 2, 4, fill_color=BG_HEADER)
    row += 1

    unit_inputs = [
        ('Average Contract Value (ACV)', ASSUMPTIONS['acv_usd'], FMT_CCY_PLAIN, 'Blended across 5 customer tiers (bank / crypto / insurance / pharma / fintech)'),
        ('Gross Margin',                  ASSUMPTIONS['gross_margin'], FMT_PCT,  'LLM API + connector infra = 22% COGS'),
        ('Customer Lifetime',             ASSUMPTIONS['customer_lifetime_yr'], '0.0" yr"', 'Mid-market compliance tools have high switching costs'),
        ('Net Revenue Retention (annual)', ASSUMPTIONS['nrr_annual'], FMT_PCT, 'Plugin upsell + connector expansion'),
        ('Customer Acquisition Cost (CAC)', ASSUMPTIONS['cac_usd'], FMT_CCY_PLAIN, 'Blended: direct $35K + channel $18K, weighted 60/40'),
        ('Lifetime Value (LTV)',          None, FMT_CCY_PLAIN, 'Derived: ACV × Lifetime × NRR'),  # derived
        ('LTV:CAC Ratio',                 None, '0.0"x"', 'Derived: LTV / CAC'),
        ('CAC Payback Period',            None, FMT_MONTHS, 'Derived: CAC / (ACV × GM / 12)'),
        ('Monthly Churn Rate',            None, FMT_PCT, 'Derived: 1 / (Lifetime × 12)'),
        ('Monthly Expansion Rate',        None, FMT_PCT, 'Derived: NRR^(1/12) − 1'),
    ]
    for label, value, fmt, note in unit_inputs:
        ws.cell(row=row, column=2, value=label).font = f_body()
        cell = ws.cell(row=row, column=3)
        cell.number_format = fmt
        if value is None:
            # Derived
            if label == 'Lifetime Value (LTV)':
                cell.value = f'=C{row-4}*C{row-3}*C{row-2}'  # ACV × Lifetime × NRR
            elif label == 'LTV:CAC Ratio':
                cell.value = f'=C{row-1}/C{row-2}'  # LTV / CAC (LTV is row-1, CAC is row-2)
            elif label == 'CAC Payback Period':
                cell.value = f'=C{row-3}/(C{row-7}*C{row-8}/12)'  # CAC / (ACV × GM / 12) — need to point to ACV and GM
            elif label == 'Monthly Churn Rate':
                cell.value = f'=1/(C{row-7}*12)'  # 1 / (Lifetime × 12)
            elif label == 'Monthly Expansion Rate':
                cell.value = f'=C{row-6}^(1/12)-1'  # NRR^(1/12) - 1 (NRR is 2 rows above)
            style_derived_cell(cell)
        else:
            cell.value = value
            style_assumption_cell(cell, note)
        ws.cell(row=row, column=4, value=note).font = f_body(size=9, italic=True, color='6B7280')
        ws.cell(row=row, column=4).alignment = align('left', 'center', wrap=True)
        ws.cell(row=row, column=2).border = border_thin()
        ws.cell(row=row, column=4).border = border_thin()
        ws.row_dimensions[row].height = 20
        row += 1

    # Track ACV row for later formulas (it was the first input row)
    acv_row = row - len(unit_inputs)  # row of ACV
    gm_row = acv_row + 1              # Gross Margin
    lifetime_row = acv_row + 2
    nrr_row = acv_row + 3
    cac_row = acv_row + 4
    ltv_row = acv_row + 5
    ratio_row = acv_row + 6
    payback_row = acv_row + 7
    monthly_churn_row = acv_row + 8
    monthly_exp_row = acv_row + 9

    # Fix the payback formula (CAC / (ACV × GM / 12))
    ws.cell(row=payback_row, column=3).value = f'=C{cac_row}/(C{acv_row}*C{gm_row}/12)'

    # Funding section
    row += 2
    ws.merge_cells(f'B{row}:D{row}')
    ws.cell(row=row, column=2, value='Funding (§5.1, §5.2)')
    style_header_row(ws, row, 2, 4, fill_color=BG_SUBHEADER)
    row += 1
    ws.cell(row=row, column=2, value='Input')
    ws.cell(row=row, column=3, value='Value')
    ws.cell(row=row, column=4, value='Source / Notes')
    style_header_row(ws, row, 2, 4, fill_color=BG_HEADER)
    row += 1

    fund_inputs = [
        ('Seed Raise',                ASSUMPTIONS['seed_raise_usd'], FMT_CCY_PLAIN, '$2.5M seed at $12M pre-money'),
        ('Target Series A ARR Trigger', ASSUMPTIONS['target_series_a_arr'], FMT_CCY_PLAIN, '$1.2M ARR — Series A gate'),
        ('Target Series A Raise',     ASSUMPTIONS['series_a_raise_usd'], FMT_CCY_PLAIN, '$10-15M target; model uses $12.5M midpoint'),
        ('Earliest Series A Month',   15, '0" mo"', 'Model waits until M15 even if ARR trigger met earlier'),
    ]
    for label, value, fmt, note in fund_inputs:
        ws.cell(row=row, column=2, value=label).font = f_body()
        cell = ws.cell(row=row, column=3, value=value)
        cell.number_format = fmt
        style_assumption_cell(cell, note)
        ws.cell(row=row, column=4, value=note).font = f_body(size=9, italic=True, color='6B7280')
        ws.cell(row=row, column=2).border = border_thin()
        ws.cell(row=row, column=4).border = border_thin()
        ws.row_dimensions[row].height = 20
        row += 1

    # Burn schedule
    row += 2
    ws.merge_cells(f'B{row}:D{row}')
    ws.cell(row=row, column=2, value='Monthly Burn Schedule (Use of Funds §5.2)')
    style_header_row(ws, row, 2, 4, fill_color=BG_SUBHEADER)
    row += 1
    ws.cell(row=row, column=2, value='Months')
    ws.cell(row=row, column=3, value='Burn ($/mo)')
    ws.cell(row=row, column=4, value='Headcount / Stage')
    style_header_row(ws, row, 2, 4, fill_color=BG_HEADER)
    row += 1

    burn_notes = [
        'Founders + 1 eng + 1 GTM (pre-seed runway)',
        '+2 eng + 1 GTM (Y1 buildout)',
        '+1 ML/infra (peak pre-A, CAC heavy)',
        'Post-A: +4 hires (2 eng, 1 sales, 1 CS)',
        'Scale GTM: +3 sales + 1 CS',
        'Full team: +2 eng + 1 ML',
    ]
    for i, ((start, end, burn), note) in enumerate(zip(ASSUMPTIONS['burn_schedule'], burn_notes)):
        ws.cell(row=row, column=2, value=f'M{start}–M{end}').font = f_body()
        cell = ws.cell(row=row, column=3, value=burn)
        cell.number_format = FMT_CCY_PLAIN
        style_assumption_cell(cell, note)
        ws.cell(row=row, column=4, value=note).font = f_body(size=9, italic=True, color='6B7280')
        ws.cell(row=row, column=2).border = border_thin()
        ws.cell(row=row, column=4).border = border_thin()
        ws.row_dimensions[row].height = 18
        row += 1

    # Scenarios section
    row += 2
    ws.merge_cells(f'B{row}:D{row}')
    ws.cell(row=row, column=2, value='Acquisition Scenarios (New Customers / Year)')
    style_header_row(ws, row, 2, 4, fill_color=BG_SUBHEADER)
    row += 1
    ws.cell(row=row, column=2, value='Scenario')
    ws.cell(row=row, column=3, value='Y1 / Y2 / Y3')
    ws.cell(row=row, column=4, value='Narrative')
    style_header_row(ws, row, 2, 4, fill_color=BG_HEADER)
    row += 1

    scenario_notes = {
        'Conservative': 'Slow design-partner ramp. 6 paid pilots in Y1, expanding only after 2 case studies published.',
        'Base':         'The plan we pitch: 12 paid customers in Y1, 28 in Y2, 48 in Y3. Assumes channel partnership signs by M9.',
        'Aggressive':   'MiCA + Consumer Duty demand pulls forward. 18 customers in Y1, 42 in Y2, 72 in Y3. Requires 2 channel partners.',
    }
    scenario_rows = {}  # for later reference
    for name, vals in ASSUMPTIONS['scenarios'].items():
        ws.cell(row=row, column=2, value=name).font = f_body(bold=True, color=vals['color'])
        cell = ws.cell(row=row, column=3, value=f"{vals['y1']} / {vals['y2']} / {vals['y3']}")
        cell.font = Font(name=FONT_BODY, size=10, color='0000FF')
        cell.fill = fill(BG_ASSUMPTION)
        cell.alignment = align('center', 'center')
        cell.border = border_thin()
        ws.cell(row=row, column=4, value=scenario_notes[name]).font = f_body(size=9, italic=True, color='6B7280')
        ws.cell(row=row, column=4).alignment = align('left', 'center', wrap=True)
        ws.cell(row=row, column=2).border = border_thin()
        ws.cell(row=row, column=4).border = border_thin()
        ws.row_dimensions[row].height = 28
        scenario_rows[name] = row
        row += 1

    return {
        'acv_row': acv_row,
        'gm_row': gm_row,
        'cac_row': cac_row,
        'lifetime_row': lifetime_row,
        'nrr_row': nrr_row,
        'monthly_churn_row': monthly_churn_row,
        'monthly_exp_row': monthly_exp_row,
        'scenario_rows': scenario_rows,
    }


# ─────────────────────────────────────────────────────────────
# Sheets 3-5: Scenario detail (Conservative / Base / Aggressive)
# ─────────────────────────────────────────────────────────────

def build_scenario_sheet(wb, scenario_name, scenario_yearly, color, refs):
    ws = wb.create_sheet(scenario_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 30
    for c in range(3, 3 + TOTAL_MONTHS + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11

    # Title
    ws.merge_cells(f'B2:{get_column_letter(2 + TOTAL_MONTHS + 1)}2')
    ws['B2'] = f'{scenario_name} Scenario — 36-Month Cohort Model'
    ws['B2'].font = Font(name=FONT_HEADING, size=14, bold=True, color=PRIMARY)
    ws.row_dimensions[2].height = 24

    ws.merge_cells(f'B3:{get_column_letter(2 + TOTAL_MONTHS + 1)}3')
    ws['B3'] = f'Acquisition: {scenario_yearly["y1"]} / {scenario_yearly["y2"]} / {scenario_yearly["y3"]} new customers per year (Y1 / Y2 / Y3)'
    ws['B3'].font = f_body(size=10, italic=True, color=color)

    # Run simulation
    months = simulate_scenario(scenario_name, scenario_yearly)

    # Header row: months M1..M36 + Total
    header_row = 5
    ws.cell(row=header_row, column=2, value='Metric')
    for i in range(TOTAL_MONTHS):
        col = 3 + i
        ws.cell(row=header_row, column=col, value=f'M{i+1}')
    ws.cell(row=header_row, column=3 + TOTAL_MONTHS, value='Total / Final')
    style_header_row(ws, header_row, 2, 3 + TOTAL_MONTHS, fill_color=BG_HEADER)

    # Section: Customer Movement
    row = header_row + 1
    section_rows = {}

    def write_row(label, values, fmt, total=None, total_fmt=None, is_total=False, color_override=None):
        nonlocal row
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=2).font = f_body(bold=is_total, color=color_override or PRIMARY)
        ws.cell(row=row, column=2).alignment = align('left', 'center')
        for i, v in enumerate(values):
            cell = ws.cell(row=row, column=3 + i, value=v)
            cell.number_format = fmt
            cell.font = f_body(bold=is_total, color=color_override or '1F2937')
            cell.alignment = align('right', 'center')
            cell.border = border_thin()
            if isinstance(v, (int, float)) and v < 0 and fmt != FMT_INT:
                cell.font = f_body(color=ACCENT_RED, bold=False)
        # Total / Final
        if total is not None:
            tcell = ws.cell(row=row, column=3 + TOTAL_MONTHS, value=total)
            tcell.number_format = total_fmt or fmt
            tcell.font = f_body(bold=True, color=PRIMARY)
            tcell.fill = fill(BG_TOTAL)
            tcell.border = border_top(color=PRIMARY, style='medium')
            tcell.alignment = align('right', 'center')
        if is_total:
            for c in range(2, 3 + TOTAL_MONTHS + 1):
                ws.cell(row=row, column=c).fill = fill(BG_TOTAL)
                ws.cell(row=row, column=c).font = f_body(bold=True, color=PRIMARY)
                ws.cell(row=row, column=c).border = border_top(color=PRIMARY, style='medium')
        ws.cell(row=row, column=2).border = border_thin()
        ws.row_dimensions[row].height = 18
        r = row
        row += 1
        return r

    # Section header: Customers
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3 + TOTAL_MONTHS)
    ws.cell(row=row, column=2, value='Customer Movement')
    style_header_row(ws, row, 2, 3 + TOTAL_MONTHS, fill_color=BG_SUBHEADER)
    row += 1

    section_rows['new_customers'] = write_row(
        'New Customers (cohort)',
        [m['new_customers'] for m in months],
        FMT_DECIMAL,
        total=sum(m['new_customers'] for m in months),
        total_fmt=FMT_INT,
    )
    section_rows['churned_customers'] = write_row(
        'Churned Customers',
        [m['churned_customers'] for m in months],
        FMT_DECIMAL,
        total=sum(m['churned_customers'] for m in months),
        total_fmt=FMT_INT,
    )
    section_rows['active_customers'] = write_row(
        'Active Customers (end of month)',
        [m['active_customers'] for m in months],
        FMT_DECIMAL,
        total=months[-1]['active_customers'],
        total_fmt=FMT_DECIMAL,
        is_total=True,
    )

    # Section header: ARR
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3 + TOTAL_MONTHS)
    ws.cell(row=row, column=2, value='ARR Movement ($)')
    style_header_row(ws, row, 2, 3 + TOTAL_MONTHS, fill_color=BG_SUBHEADER)
    row += 1

    section_rows['new_arr'] = write_row(
        'New ARR',
        [m['new_arr'] for m in months],
        FMT_CCY_PLAIN,
        total=sum(m['new_arr'] for m in months),
        total_fmt=FMT_CCY_PLAIN,
        color_override=ACCENT_GREEN,
    )
    section_rows['expansion_arr'] = write_row(
        'Expansion ARR',
        [m['expansion_arr'] for m in months],
        FMT_CCY_PLAIN,
        total=sum(m['expansion_arr'] for m in months),
        total_fmt=FMT_CCY_PLAIN,
        color_override=ACCENT_GREEN,
    )
    section_rows['churned_arr'] = write_row(
        'Churned ARR',
        [-m['churned_arr'] for m in months],
        FMT_CCY_PLAIN,
        total=-sum(m['churned_arr'] for m in months),
        total_fmt=FMT_CCY_PLAIN,
        color_override=ACCENT_RED,
    )
    section_rows['net_new_arr'] = write_row(
        'Net New ARR',
        [m['net_new_arr'] for m in months],
        FMT_CCY_PLAIN,
        total=sum(m['net_new_arr'] for m in months),
        total_fmt=FMT_CCY_PLAIN,
        is_total=True,
    )
    section_rows['cumulative_arr'] = write_row(
        'Cumulative ARR',
        [m['cumulative_arr'] for m in months],
        FMT_CCY_PLAIN,
        total=months[-1]['cumulative_arr'],
        total_fmt=FMT_CCY_PLAIN,
        is_total=True,
    )
    # ARR as $M for readability
    section_rows['cumulative_arr_m'] = write_row(
        'Cumulative ARR ($M)',
        [m['cumulative_arr'] / 1_000_000 for m in months],
        FMT_CCY_M,
        total=months[-1]['cumulative_arr'] / 1_000_000,
        total_fmt=FMT_CCY_M,
    )

    # Section header: Cash Flow
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3 + TOTAL_MONTHS)
    ws.cell(row=row, column=2, value='Cash Flow')
    style_header_row(ws, row, 2, 3 + TOTAL_MONTHS, fill_color=BG_SUBHEADER)
    row += 1

    section_rows['gross_profit'] = write_row(
        'Gross Profit (ARR/12 × GM)',
        [m['gross_profit'] for m in months],
        FMT_CCY_PLAIN,
        total=sum(m['gross_profit'] for m in months),
        total_fmt=FMT_CCY_PLAIN,
    )
    section_rows['burn'] = write_row(
        'Monthly Burn',
        [-m['burn'] for m in months],
        FMT_CCY_PLAIN,
        total=-sum(m['burn'] for m in months),
        total_fmt=FMT_CCY_PLAIN,
        color_override=ACCENT_RED,
    )
    section_rows['series_a'] = write_row(
        'Series A Cash Inflow',
        [m['series_a_cash'] for m in months],
        FMT_CCY_PLAIN,
        total=sum(m['series_a_cash'] for m in months),
        total_fmt=FMT_CCY_PLAIN,
        color_override=ACCENT_GREEN,
    )
    section_rows['net_cash_flow'] = write_row(
        'Net Cash Flow',
        [m['net_cash_flow'] for m in months],
        FMT_CCY_PLAIN,
        total=sum(m['net_cash_flow'] for m in months),
        total_fmt=FMT_CCY_PLAIN,
        is_total=True,
    )
    section_rows['cash_balance'] = write_row(
        'Cash Balance (end of month)',
        [m['cash_balance'] for m in months],
        FMT_CCY_PLAIN,
        total=months[-1]['cash_balance'],
        total_fmt=FMT_CCY_PLAIN,
        is_total=True,
    )

    # Runway row
    runway_values = []
    for m in months:
        if m['runway_months'] == float('inf'):
            runway_values.append('Profitable')
        else:
            runway_values.append(round(m['runway_months'], 1))
    runway_cell_row = row
    ws.cell(row=row, column=2, value='Runway (months at current burn)')
    ws.cell(row=row, column=2).font = f_body(bold=True, color=PRIMARY)
    ws.cell(row=row, column=2).alignment = align('left', 'center')
    ws.cell(row=row, column=2).border = border_thin()
    for i, v in enumerate(runway_values):
        cell = ws.cell(row=row, column=3 + i, value=v)
        cell.font = f_body(bold=True, color=ACCENT_GREEN if v == 'Profitable' else PRIMARY)
        cell.alignment = align('right', 'center')
        cell.border = border_thin()
        if isinstance(v, str):
            cell.number_format = '@'
        else:
            cell.number_format = FMT_MONTHS
            if v < 6:
                cell.font = f_body(bold=True, color=ACCENT_RED)
            elif v < 12:
                cell.font = f_body(bold=True, color=ACCENT_GOLD)
    # Final runway
    final_runway = runway_values[-1]
    fcell = ws.cell(row=row, column=3 + TOTAL_MONTHS, value=final_runway if isinstance(final_runway, str) else round(final_runway, 1))
    fcell.font = f_body(bold=True, color=PRIMARY)
    fcell.fill = fill(BG_TOTAL)
    fcell.border = border_top(color=PRIMARY, style='medium')
    fcell.alignment = align('right', 'center')
    if isinstance(final_runway, str):
        fcell.number_format = '@'
    else:
        fcell.number_format = FMT_MONTHS
    ws.row_dimensions[row].height = 18
    row += 1

    # Series A trigger row
    sa_row = row
    ws.cell(row=row, column=2, value='Series A Trigger')
    ws.cell(row=row, column=2).font = f_body(bold=True, color=PRIMARY)
    ws.cell(row=row, column=2).alignment = align('left', 'center')
    ws.cell(row=row, column=2).border = border_thin()
    for i, m in enumerate(months):
        cell = ws.cell(row=row, column=3 + i, value='▲ Triggered' if m['series_a_received'] and m['series_a_cash'] > 0 else ('✓ Closed' if m['series_a_received'] else ''))
        cell.font = f_body(color=ACCENT_GREEN if m['series_a_received'] else '6B7280', bold=m['series_a_received'] and m['series_a_cash'] > 0)
        cell.alignment = align('center', 'center')
        cell.border = border_thin()
    ws.cell(row=row, column=3 + TOTAL_MONTHS, value='Triggered' if months[-1]['series_a_received'] else 'NOT triggered')
    ws.cell(row=row, column=3 + TOTAL_MONTHS).font = f_body(bold=True, color=ACCENT_GREEN if months[-1]['series_a_received'] else ACCENT_RED)
    ws.cell(row=row, column=3 + TOTAL_MONTHS).fill = fill(BG_TOTAL)
    ws.cell(row=row, column=3 + TOTAL_MONTHS).border = border_top(color=PRIMARY, style='medium')
    ws.cell(row=row, column=3 + TOTAL_MONTHS).alignment = align('center', 'center')
    ws.row_dimensions[row].height = 18
    row += 1

    # ─── Chart: ARR growth + Cash balance ───
    chart_anchor_row = row + 2
    chart = LineChart()
    chart.title = f'{scenario_name} — Cumulative ARR + Cash Balance'
    chart.style = 2
    chart.y_axis.title = 'USD'
    chart.x_axis.title = 'Month'
    chart.height = 10
    chart.width = 22

    # ARR data (Cumulative ARR row, columns C..AM)
    arr_data = Reference(ws, min_col=2, min_row=section_rows['cumulative_arr'],
                         max_col=2 + TOTAL_MONTHS, max_row=section_rows['cumulative_arr'])
    chart.add_data(arr_data, titles_from_data=True, from_rows=True)

    cash_data = Reference(ws, min_col=2, min_row=section_rows['cash_balance'],
                          max_col=2 + TOTAL_MONTHS, max_row=section_rows['cash_balance'])
    chart.add_data(cash_data, titles_from_data=True, from_rows=True)

    # Categories: M1..M36
    cats = Reference(ws, min_col=3, min_row=header_row,
                     max_col=2 + TOTAL_MONTHS, max_row=header_row)
    chart.set_categories(cats)

    ws.add_chart(chart, f'B{chart_anchor_row}')

    # Freeze panes
    ws.freeze_panes = 'C6'

    return months, section_rows


# ─────────────────────────────────────────────────────────────
# Sheet 6: Summary — 3-scenario comparison
# ─────────────────────────────────────────────────────────────

def build_summary(wb, scenarios_data):
    """scenarios_data: list of (name, color, months, section_rows)"""
    ws = wb.create_sheet('Summary')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 38
    for i, (name, *_rest) in enumerate(scenarios_data):
        ws.column_dimensions[get_column_letter(3 + i)].width = 20

    # Title
    ws['B2'] = 'Scenario Summary — Side-by-Side Comparison'
    ws['B2'].font = Font(name=FONT_HEADING, size=16, bold=True, color=PRIMARY)
    ws.row_dimensions[2].height = 26
    ws['B3'] = 'Compares Conservative / Base / Aggressive across ARR, cash, runway, and Series A timing.'
    ws['B3'].font = f_body(size=10, italic=True, color='6B7280')

    # Headers
    header_row = 5
    ws.cell(row=header_row, column=2, value='Metric')
    for i, (name, color, *_rest) in enumerate(scenarios_data):
        ws.cell(row=header_row, column=3 + i, value=name)
    style_header_row(ws, header_row, 2, 2 + len(scenarios_data))

    # Helper to write a metric row
    row = header_row + 1
    def metric_row(label, values, fmt, color_override=None, bold=False):
        nonlocal row
        ws.cell(row=row, column=2, value=label).font = f_body(bold=bold, color=PRIMARY)
        ws.cell(row=row, column=2).alignment = align('left', 'center')
        ws.cell(row=row, column=2).border = border_thin()
        for i, v in enumerate(values):
            cell = ws.cell(row=row, column=3 + i, value=v)
            cell.number_format = fmt
            cell.font = f_body(bold=bold, color=color_override or '1F2937')
            cell.alignment = align('right', 'center')
            cell.border = border_thin()
        ws.row_dimensions[row].height = 18
        row += 1

    # Section header: ARR
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + len(scenarios_data))
    ws.cell(row=row, column=2, value='ARR Trajectory')
    style_header_row(ws, row, 2, 2 + len(scenarios_data), fill_color=BG_SUBHEADER)
    row += 1

    metric_row('ARR at M12',
               [next(m['cumulative_arr'] for m in months if m['month'] == 12) for (_, _, months, _) in scenarios_data],
               FMT_CCY_PLAIN)
    metric_row('ARR at M18 (Series A trigger)',
               [next(m['cumulative_arr'] for m in months if m['month'] == 18) for (_, _, months, _) in scenarios_data],
               FMT_CCY_PLAIN, bold=True)
    metric_row('ARR at M24',
               [next(m['cumulative_arr'] for m in months if m['month'] == 24) for (_, _, months, _) in scenarios_data],
               FMT_CCY_PLAIN)
    metric_row('ARR at M36',
               [next(m['cumulative_arr'] for m in months if m['month'] == 36) for (_, _, months, _) in scenarios_data],
               FMT_CCY_PLAIN, bold=True, color_override=ACCENT_GREEN)
    metric_row('ARR at M36 ($M)',
               [next(m['cumulative_arr'] for m in months if m['month'] == 36) / 1_000_000 for (_, _, months, _) in scenarios_data],
               FMT_CCY_M)

    # Section header: Customers
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + len(scenarios_data))
    ws.cell(row=row, column=2, value='Customer Counts')
    style_header_row(ws, row, 2, 2 + len(scenarios_data), fill_color=BG_SUBHEADER)
    row += 1

    metric_row('Active Customers at M12',
               [round(next(m['active_customers'] for m in months if m['month'] == 12), 1) for (_, _, months, _) in scenarios_data],
               FMT_DECIMAL)
    metric_row('Active Customers at M24',
               [round(next(m['active_customers'] for m in months if m['month'] == 24), 1) for (_, _, months, _) in scenarios_data],
               FMT_DECIMAL)
    metric_row('Active Customers at M36',
               [round(next(m['active_customers'] for m in months if m['month'] == 36), 1) for (_, _, months, _) in scenarios_data],
               FMT_DECIMAL, bold=True)
    metric_row('Cumulative New Customers (36mo)',
               [round(sum(m['new_customers'] for m in months), 0) for (_, _, months, _) in scenarios_data],
               FMT_INT)
    metric_row('Cumulative Churned Customers (36mo)',
               [round(sum(m['churned_customers'] for m in months), 0) for (_, _, months, _) in scenarios_data],
               FMT_INT, color_override=ACCENT_RED)

    # Section header: Cash & Runway
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + len(scenarios_data))
    ws.cell(row=row, column=2, value='Cash & Runway')
    style_header_row(ws, row, 2, 2 + len(scenarios_data), fill_color=BG_SUBHEADER)
    row += 1

    metric_row('Peak Monthly Burn',
               [max(m['burn'] for m in months) for (_, _, months, _) in scenarios_data],
               FMT_CCY_PLAIN, color_override=ACCENT_RED)
    metric_row('Total Burn (36mo)',
               [sum(m['burn'] for m in months) for (_, _, months, _) in scenarios_data],
               FMT_CCY_PLAIN, color_override=ACCENT_RED)
    metric_row('Cash Balance at M18',
               [next(m['cash_balance'] for m in months if m['month'] == 18) for (_, _, months, _) in scenarios_data],
               FMT_CCY_PLAIN, bold=True)
    metric_row('Cash Balance at M24',
               [next(m['cash_balance'] for m in months if m['month'] == 24) for (_, _, months, _) in scenarios_data],
               FMT_CCY_PLAIN)
    metric_row('Cash Balance at M36',
               [next(m['cash_balance'] for m in months if m['month'] == 36) for (_, _, months, _) in scenarios_data],
               FMT_CCY_PLAIN, bold=True, color_override=ACCENT_GREEN)

    # Runway at M18 (pre-A trigger)
    runway_at_m18 = []
    for (_, _, months, _) in scenarios_data:
        m18 = next(m for m in months if m['month'] == 18)
        if m18['runway_months'] == float('inf'):
            runway_at_m18.append('Profitable')
        else:
            runway_at_m18.append(round(m18['runway_months'], 1))
    ws.cell(row=row, column=2, value='Runway at M18 (pre-A)').font = f_body(bold=True, color=PRIMARY)
    ws.cell(row=row, column=2).border = border_thin()
    for i, v in enumerate(runway_at_m18):
        cell = ws.cell(row=row, column=3 + i, value=v)
        if isinstance(v, str):
            cell.number_format = '@'
            cell.font = f_body(color=ACCENT_GREEN, bold=True)
        else:
            cell.number_format = FMT_MONTHS
            cell.font = f_body(color=ACCENT_RED if v < 6 else (ACCENT_GOLD if v < 12 else ACCENT_GREEN), bold=True)
        cell.alignment = align('right', 'center')
        cell.border = border_thin()
    ws.row_dimensions[row].height = 18
    row += 1

    # Runway at M36
    runway_at_m36 = []
    for (_, _, months, _) in scenarios_data:
        m36 = months[-1]
        if m36['runway_months'] == float('inf'):
            runway_at_m36.append('Profitable')
        else:
            runway_at_m36.append(round(m36['runway_months'], 1))
    ws.cell(row=row, column=2, value='Runway at M36').font = f_body(bold=True, color=PRIMARY)
    ws.cell(row=row, column=2).border = border_thin()
    for i, v in enumerate(runway_at_m36):
        cell = ws.cell(row=row, column=3 + i, value=v)
        if isinstance(v, str):
            cell.number_format = '@'
            cell.font = f_body(color=ACCENT_GREEN, bold=True)
        else:
            cell.number_format = FMT_MONTHS
            cell.font = f_body(color=ACCENT_RED if v < 6 else (ACCENT_GOLD if v < 12 else ACCENT_GREEN), bold=True)
        cell.alignment = align('right', 'center')
        cell.border = border_thin()
    ws.row_dimensions[row].height = 18
    row += 1

    # Lowest cash point (trough)
    troughs = []
    trough_months = []
    for (_, _, months, _) in scenarios_data:
        min_cash = min(m['cash_balance'] for m in months)
        trough_month = next(m['month'] for m in months if m['cash_balance'] == min_cash)
        troughs.append(min_cash)
        trough_months.append(trough_month)
    metric_row('Lowest Cash Point (trough)',
               troughs, FMT_CCY_PLAIN, color_override=ACCENT_RED, bold=True)
    ws.cell(row=row - 1, column=2).value = f'Lowest Cash Point (trough, month)'
    # Replace trough month column with text
    for i, tm in enumerate(trough_months):
        cell = ws.cell(row=row - 1, column=3 + i, value=f'M{tm}')
        cell.font = f_body(color=ACCENT_RED, bold=True)
        cell.alignment = align('right', 'center')
        cell.number_format = '@'

    # Section header: Series A
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + len(scenarios_data))
    ws.cell(row=row, column=2, value='Series A Timing')
    style_header_row(ws, row, 2, 2 + len(scenarios_data), fill_color=BG_SUBHEADER)
    row += 1

    # Series A month (first month where series_a_cash > 0)
    sa_months = []
    for (_, _, months, _) in scenarios_data:
        sa_month = next((m['month'] for m in months if m['series_a_cash'] > 0), None)
        sa_months.append(sa_month)
    ws.cell(row=row, column=2, value='Series A Close Month').font = f_body(bold=True, color=PRIMARY)
    ws.cell(row=row, column=2).border = border_thin()
    for i, m in enumerate(sa_months):
        cell = ws.cell(row=row, column=3 + i, value=f'M{m}' if m else 'NOT triggered')
        cell.font = f_body(color=ACCENT_GREEN if m else ACCENT_RED, bold=True)
        cell.alignment = align('right', 'center')
        cell.number_format = '@'
        cell.border = border_thin()
    ws.row_dimensions[row].height = 18
    row += 1

    metric_row('ARR at Series A Close',
               [next(m['cumulative_arr'] for m in months if m['month'] == sa_months[i]) if sa_months[i] else 0
                for i, (_, _, months, _) in enumerate(scenarios_data)],
               FMT_CCY_PLAIN, color_override=ACCENT_GREEN, bold=True)

    # Section header: Investor Proof Points
    row += 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + len(scenarios_data))
    ws.cell(row=row, column=2, value='Investor Proof Points')
    style_header_row(ws, row, 2, 2 + len(scenarios_data), fill_color=BG_SUBHEADER)
    row += 1

    # LTV:CAC consistency check
    metric_row('LTV:CAC Ratio (steady-state)',
               [ASSUMPTIONS['ltv_cac_ratio']] * len(scenarios_data),
               '0.0"x"', color_override=ACCENT_GREEN, bold=True)
    metric_row('CAC Payback (months, steady-state)',
               [ASSUMPTIONS['cac_payback_months']] * len(scenarios_data),
               FMT_MONTHS, color_override=ACCENT_GREEN, bold=True)
    metric_row('Gross Margin (steady-state)',
               [ASSUMPTIONS['gross_margin']] * len(scenarios_data),
               FMT_PCT, color_override=ACCENT_GREEN, bold=True)

    # Footnote
    row += 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + len(scenarios_data))
    ws.cell(row=row, column=2, value='Note: LTV:CAC, CAC Payback, and Gross Margin are steady-state unit economics from §4.2. They do not vary by scenario because they depend on the per-customer profile, not the count of customers.')
    ws.cell(row=row, column=2).font = f_body(size=9, italic=True, color='6B7280')
    ws.cell(row=row, column=2).alignment = align('left', 'top', wrap=True)
    ws.row_dimensions[row].height = 30

    # ─── Chart: ARR comparison across 3 scenarios ───
    chart_row = row + 3
    chart = LineChart()
    chart.title = 'Cumulative ARR — Scenario Comparison'
    chart.style = 2
    chart.y_axis.title = 'ARR ($)'
    chart.x_axis.title = 'Month'
    chart.height = 12
    chart.width = 22

    for i, (name, color, months, srows) in enumerate(scenarios_data):
        # Reference the cumulative_arr row on each scenario sheet
        # We need a 1-based col index = 2 (label) + 1..36 (months)
        arr_row = srows['cumulative_arr']
        data = Reference(wb[name], min_col=2, min_row=arr_row,
                         max_col=2 + TOTAL_MONTHS, max_row=arr_row)
        chart.add_data(data, titles_from_data=True, from_rows=True)

    # Categories: M1..M36 from the first scenario sheet
    first_sheet_name = scenarios_data[0][0]
    cats = Reference(wb[first_sheet_name], min_col=3, min_row=5,
                     max_col=2 + TOTAL_MONTHS, max_row=5)
    chart.set_categories(cats)

    ws.add_chart(chart, f'B{chart_row}')


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────

def main():
    output_path = Path('/home/z/my-project/download/RegGuard_AI_36Month_Cohort_Model.xlsx')
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default sheet (we'll add our own)
    wb.remove(wb.active)

    # 1. README
    build_readme(wb)

    # 2. Assumptions
    refs = build_assumptions(wb)

    # 3-5. Scenario sheets
    scenarios_data = []
    for name, vals in ASSUMPTIONS['scenarios'].items():
        months, srows = build_scenario_sheet(wb, name, vals, vals['color'], refs)
        scenarios_data.append((name, vals['color'], months, srows))

    # 6. Summary
    build_summary(wb, scenarios_data)

    # Metadata
    wb.properties.creator = 'Z.ai — RegGuard AI'
    wb.properties.title = 'RegGuard AI — 36-Month Cohort Financial Model'
    wb.properties.subject = 'ARR growth, burn, runway under 3 acquisition scenarios'

    # Save
    wb.save(output_path)
    print(f'✓ Wrote {output_path}')
    print(f'  Sheets: {wb.sheetnames}')

    # Print scenario summary
    print()
    print('Scenario Summary (Month 36):')
    print(f'  {"Scenario":<14} {"ARR":>15} {"Customers":>12} {"Cash":>15} {"Series A":>10}')
    for name, color, months, srows in scenarios_data:
        m36 = months[-1]
        sa_month = next((m['month'] for m in months if m['series_a_cash'] > 0), None)
        sa_str = f'M{sa_month}' if sa_month else 'NOT TRIG'
        print(f'  {name:<14} ${m36["cumulative_arr"]:>13,} {m36["active_customers"]:>12.1f} ${m36["cash_balance"]:>13,} {sa_str:>10}')


if __name__ == '__main__':
    main()
